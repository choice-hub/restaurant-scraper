"""
Bolt Food restaurant scraper — Playwright browser automation.

Strategy:
- Navigate directly to food.bolt.eu/en/{city_id}-{city_slug}/
- Extract all data from aria-label and href on card buttons (no detail pages needed
  for the core fields)
- --detail flag visits each restaurant page to add cuisine + address

Usage:
  python bolt_scraper.py --city "Brno" --max 200
  python bolt_scraper.py --city "Prague" --max 500 --detail
  python bolt_scraper.py --city "Warsaw" --max 100 --headless
"""
import argparse
import asyncio
import csv
import json
import logging
import os
import random
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

# ── Known city IDs (found by running with --discover) ─────────────────────
# Format: lowercase city name → (city_id, url_slug)
KNOWN_CITIES: dict[str, tuple[int, str]] = {
    "brno":     (456, "brno"),
    "prague":   (271, "prague"),
    "bratislava": (270, "bratislava"),
    "warsaw":   (300, "warsaw"),
    "riga":     (302, "riga"),
    "tallinn":  (303, "tallinn"),
    "vilnius":  (301, "vilnius"),
}

BRNO_LAT,  BRNO_LNG  = 49.1950602, 16.6068371   # fallback coords for unknown city discovery
IP_LAT_PRAGUE, IP_LNG_PRAGUE = 50.0805, 14.467   # Bolt's default IP location (Prague)

# ── Config ─────────────────────────────────────────────────────────────────
CHECKPOINT_EVERY  = 15
SCROLL_PAUSE      = (1.5, 3.5)
ACTION_PAUSE      = (1.0, 2.5)
MAX_STALLS        = 8
PAGE_LOAD_WAIT    = 25_000   # ms
DETAIL_TIMEOUT    = 15_000

LOG_DIR = Path("bolt_logs")
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(LOG_DIR / f"scrape_{datetime.now():%Y%m%d_%H%M%S}.log"),
    ],
)
log = logging.getLogger(__name__)

# ── Aria-label parsing ─────────────────────────────────────────────────────
# Example: "KFC Brno Náměstí Svobody, Delivery price 18 Kč, Delivery time 10 to 15 minutes, Rating 4.4"
_RE_PRICE  = re.compile(r"Delivery price\s+(.+?)(?:,\s*Delivery time|$)", re.I)
_RE_TIME   = re.compile(r"Delivery time\s+(.+?)(?:,\s*Rating|$)", re.I)
_RE_RATING = re.compile(r"Rating\s+([\d.,]+)", re.I)
_RE_REVIEW = re.compile(r"\(([\d+,]+\+?)\)")  # (500+) or (1,234)

def parse_aria(label: str) -> dict:
    """Extract structured data from a ProviderCard button's aria-label."""
    # Name is everything before the first ", Delivery"
    name_match = re.match(r"^(.+?)(?:,\s*Delivery price|$)", label, re.I)
    name = name_match.group(1).strip() if name_match else label.strip()

    m_price  = _RE_PRICE.search(label)
    m_time   = _RE_TIME.search(label)
    m_rating = _RE_RATING.search(label)

    return {
        "restaurant_name": name,
        "delivery_fee":    m_price.group(1).strip()  if m_price  else "",
        "delivery_time":   m_time.group(1).strip()   if m_time   else "",
        "rating":          m_rating.group(1).strip() if m_rating else "",
    }


# ── Utils ──────────────────────────────────────────────────────────────────
async def rand_sleep(lo, hi):
    await asyncio.sleep(random.uniform(lo, hi))

async def screenshot(page, label):
    try:
        p = LOG_DIR / f"shot_{label}_{datetime.now():%H%M%S}.png"
        await page.screenshot(path=str(p), full_page=False)
        log.info(f"Screenshot → {p.name}")
    except Exception:
        pass

def save_csv(data, path):
    if not data: return
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=data[0].keys())
        w.writeheader(); w.writerows(data)
    log.info(f"CSV → {path}")

def save_xlsx(data, path):
    if not data: return
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Bolt Food"
    headers = list(data[0].keys())
    hfont = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="34C759")
    halign = Alignment(horizontal="center", vertical="center")
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment = hfont, hfill, halign
    for ri, row in enumerate(data, 2):
        for ci, key in enumerate(headers, 1):
            ws.cell(row=ri, column=ci, value=row.get(key, ""))
    for ci in range(1, len(headers) + 1):
        cl = get_column_letter(ci)
        mx = max(len(str(ws.cell(row=r, column=ci).value or "")) for r in range(1, min(len(data)+2, 200)))
        ws.column_dimensions[cl].width = min(mx + 4, 55)
    wb.save(path)
    log.info(f"Excel → {path}")


# ── City ID discovery ──────────────────────────────────────────────────────
async def discover_city_id(city: str, city_lat: float, city_lng: float, headless: bool):
    """
    Open food.bolt.eu, mock the IP location to city coords, type city name,
    click first matching suggestion, and extract city_id + slug from resulting URL.
    Returns (city_id, slug) or None on failure.
    """
    log.info(f"Discovering city ID for '{city}' ...")

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=headless,
            args=["--disable-blink-features=AutomationControlled"],
        )
        ctx = await browser.new_context(
            viewport={"width": 1280, "height": 900},
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
        )
        page = await ctx.new_page()

        # Mock suggestDeliveryLocations to use city coordinates instead of Prague IP
        async def route_suggest(route):
            url = route.request.url
            if "suggestDeliveryLocations" in url:
                # Replace Prague IP coords with city coords
                url = (url
                    .replace(f"lat={IP_LAT_PRAGUE}", f"lat={city_lat}")
                    .replace(f"lng={IP_LNG_PRAGUE}", f"lng={city_lng}")
                    .replace("lat=50.0805", f"lat={city_lat}")
                    .replace("lng=14.467", f"lng={city_lng}")
                )
                import urllib.request as ur
                try:
                    req = ur.Request(url, headers={
                        k: v for k, v in dict(route.request.headers).items()
                        if k.lower() not in ('host', 'content-length')
                    })
                    with ur.urlopen(req, timeout=10) as resp:
                        await route.fulfill(status=200, body=resp.read(), content_type="application/json")
                        return
                except Exception as e:
                    log.warning(f"Route mock failed: {e}")
            await route.continue_()

        await page.route("**/*", route_suggest)
        await page.goto("https://food.bolt.eu", timeout=30_000, wait_until="domcontentloaded")
        await rand_sleep(2, 3)

        # Accept cookies
        for text in ["Allow all", "Accept all"]:
            try:
                btn = page.locator(f"button:has-text('{text}')").first
                if await btn.is_visible(timeout=3_000):
                    await btn.click(); await rand_sleep(0.8, 1.5); break
            except Exception: pass

        # Click input container
        try:
            container = page.locator("[data-testid='screens.DestinationLanding.searchInput.container']").first
            await container.click(timeout=5_000)
            await rand_sleep(0.5, 1.0)
        except Exception: pass

        # Type city name
        inp = page.locator("input[placeholder*='address' i]").first
        await inp.type(city, delay=random.randint(60, 100))
        await rand_sleep(3, 5)  # wait for suggestions

        # Pick first suggestion containing the city name
        sugg_sel = "[data-testid='screens.DestinationLanding.searchSuggestion']"
        try:
            await page.wait_for_selector(sugg_sel, timeout=10_000)
        except PWTimeout:
            log.error(f"No suggestions for '{city}'"); await browser.close(); return None

        items = await page.locator(sugg_sel).all()
        log.info(f"Suggestions: {[((await i.inner_text()).strip()) for i in items[:5]]}")

        # Pick suggestion containing city keyword (case-insensitive)
        target = None
        for item in items:
            try:
                t = (await item.inner_text(timeout=1_000)).strip()
                if city.lower() in t.lower():
                    target = item; break
            except Exception: pass
        if not target and items:
            target = items[0]

        if not target:
            log.error("No suggestion to click"); await browser.close(); return None

        await target.click()
        await rand_sleep(2, 4)

        url = page.url
        log.info(f"URL after selection: {url}")
        await browser.close()

    # Parse /en/{id}-{slug}/
    m = re.search(r"/en/(\d+)-([^/?#]+)", url)
    if m:
        city_id = int(m.group(1))
        slug = m.group(2)
        log.info(f"Discovered: city_id={city_id}, slug={slug!r}")
        return city_id, slug
    log.error(f"Could not parse city_id from URL: {url}")
    return None


# ── Browser context factory ────────────────────────────────────────────────
async def make_context(pw, headless: bool):
    browser = await pw.chromium.launch(
        headless=headless,
        args=["--disable-blink-features=AutomationControlled"],
    )
    ctx = await browser.new_context(
        viewport={"width": 1280, "height": 900},
        user_agent=(
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        locale="en-US",
    )
    return browser, ctx


# ── Detail page scrape ────────────────────────────────────────────────────
async def scrape_detail(page, url: str) -> dict:
    """Visit restaurant page, extract cuisine categories and address."""
    extra = {"cuisine": "", "address": ""}
    if not url:
        return extra
    try:
        await page.goto(url, timeout=DETAIL_TIMEOUT, wait_until="domcontentloaded")
        await rand_sleep(1.5, 3.0)

        # Cuisine from category tags / breadcrumbs
        for sel in [
            "[data-testid*='category' i]",
            "[data-testid*='cuisine' i]",
            "[class*='category' i]",
            "[class*='tag' i]",
        ]:
            try:
                items = await page.locator(sel).all()
                texts = []
                for item in items[:5]:
                    t = (await item.inner_text(timeout=1_000)).strip()
                    if t and len(t) < 60:
                        texts.append(t)
                if texts:
                    extra["cuisine"] = ", ".join(texts)
                    break
            except Exception:
                pass

        # Address
        for sel in ["[data-testid*='address' i]", "[itemprop='address']", "address"]:
            try:
                t = (await page.locator(sel).first.inner_text(timeout=2_000)).strip()
                if t:
                    extra["address"] = t
                    break
            except Exception:
                pass

    except Exception as e:
        log.warning(f"Detail page error {url}: {e}")
    return extra


# ── Main scraping loop ─────────────────────────────────────────────────────
CARD_SEL    = "[data-testid='components.ProviderCard.horizontalView']"
RATING_SEL  = "[data-testid='components.ProviderCard.providerRatingBadge']"


async def scrape_listing(
    page, city: str, city_url: str, max_count: int,
    checkpoint_path: Path, use_detail: bool, detail_page
) -> list[dict]:

    await page.goto(city_url, timeout=30_000, wait_until="domcontentloaded")
    await rand_sleep(2, 3)

    # Accept cookies
    for text in ["Allow all", "Accept all", "Reject all"]:
        try:
            btn = page.locator(f"button:has-text('{text}')").first
            if await btn.is_visible(timeout=3_000):
                await btn.click(); await rand_sleep(0.8, 1.5); break
        except Exception: pass

    await screenshot(page, "listing")

    # Wait for first cards
    try:
        await page.wait_for_selector(CARD_SEL, timeout=PAGE_LOAD_WAIT)
    except PWTimeout:
        log.error("No restaurant cards found on listing page")
        await screenshot(page, "no_cards")
        return []

    timestamp = datetime.now().isoformat()
    seen: set = set()
    results: list = []

    # Resume from checkpoint
    if checkpoint_path.exists():
        with open(checkpoint_path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                results.append(dict(row))
                seen.add(row.get("restaurant_url") or row.get("restaurant_name", ""))
        log.info(f"Resumed: {len(results)} already collected")

    stalls = 0
    prev_count = 0

    while len(results) < max_count:
        # Count current cards
        cards = page.locator(CARD_SEL)
        total = await cards.count()

        if total == prev_count:
            stalls += 1
            log.info(f"Stall {stalls}/{MAX_STALLS} | cards={total}")
            if stalls >= MAX_STALLS:
                log.info("Reached end of listing")
                break
            await page.evaluate("window.scrollBy(0, window.innerHeight * 0.7)")
            await rand_sleep(*SCROLL_PAUSE)
            continue

        stalls = 0

        # Process new cards since last scroll
        for i in range(prev_count, total):
            if len(results) >= max_count:
                break
            try:
                card = cards.nth(i)

                # Get the button with href and aria-label
                btn = card.locator("button[href]").first
                aria = await btn.get_attribute("aria-label", timeout=2_000) or ""
                href = await btn.get_attribute("href", timeout=2_000) or ""

                if not aria:
                    continue

                rec = parse_aria(aria)
                restaurant_url = f"https://food.bolt.eu{href}" if href.startswith("/") else href

                # Dedup
                key = restaurant_url or rec["restaurant_name"]
                if not key or key in seen:
                    continue
                seen.add(key)

                # Review count from rating badge text
                review_count = ""
                try:
                    badge_text = (await card.locator(RATING_SEL).first.inner_text(timeout=1_500)).strip()
                    m = _RE_REVIEW.search(badge_text)
                    if m:
                        review_count = m.group(1)
                except Exception:
                    pass

                rec["review_count"]    = review_count
                rec["restaurant_url"]  = restaurant_url
                rec["city"]            = city
                rec["cuisine"]         = ""
                rec["address"]         = ""
                rec["source_timestamp"] = timestamp

                # Optional detail page
                if use_detail and detail_page and restaurant_url:
                    extra = await scrape_detail(detail_page, restaurant_url)
                    rec.update(extra)

                results.append(rec)
                log.info(
                    f"[{len(results)}] {rec['restaurant_name']} | "
                    f"⭐{rec['rating']} ({rec['review_count']}) | "
                    f"{rec['delivery_time']} | {rec['delivery_fee']}"
                )

                if len(results) % CHECKPOINT_EVERY == 0:
                    save_csv(results, checkpoint_path)
                    log.info(f"Checkpoint: {len(results)}")

            except Exception as e:
                log.debug(f"Card {i} error: {e}")

        prev_count = total

        # Scroll down for more
        await page.evaluate("window.scrollBy(0, window.innerHeight * 0.9)")
        await rand_sleep(*SCROLL_PAUSE)

    return results


# ── Entry point ───────────────────────────────────────────────────────────
async def run(city: str, max_count: int, headless: bool, use_detail: bool,
              city_lat: float, city_lng: float):

    # Resolve city URL
    city_key = city.strip().lower()
    if city_key in KNOWN_CITIES:
        city_id, slug = KNOWN_CITIES[city_key]
        log.info(f"Using known city: {city} → city_id={city_id}, slug={slug}")
    else:
        result = await discover_city_id(city, city_lat, city_lng, headless)
        if not result:
            log.error(f"Could not find Bolt Food city ID for '{city}'. "
                      "Add it to KNOWN_CITIES or provide correct coordinates with --lat/--lng.")
            return
        city_id, slug = result
        log.info(f"Add to KNOWN_CITIES: '{city_key}': ({city_id}, '{slug}')")

    city_url = f"https://food.bolt.eu/en/{city_id}-{slug}/"
    log.info(f"Listing URL: {city_url}")

    safe_city = city.replace(" ", "_")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    checkpoint_path = LOG_DIR / f"checkpoint_{safe_city}.csv"
    out_csv  = Path(f"bolt_{safe_city}_{ts}.csv")
    out_xlsx = Path(f"bolt_{safe_city}_{ts}.xlsx")

    async with async_playwright() as pw:
        browser, ctx = await make_context(pw, headless)
        page = await ctx.new_page()

        detail_page = None
        if use_detail:
            detail_page = await ctx.new_page()

        try:
            results = await scrape_listing(
                page, city, city_url, max_count,
                checkpoint_path, use_detail, detail_page
            )
        finally:
            if detail_page:
                await detail_page.close()
            await browser.close()

    if not results:
        log.error("No restaurants collected.")
        return

    # Reorder columns for output
    ordered_fields = [
        "restaurant_name", "cuisine", "rating", "review_count",
        "delivery_fee", "delivery_time", "address",
        "restaurant_url", "city", "source_timestamp",
    ]
    results = [{k: r.get(k, "") for k in ordered_fields} for r in results]

    save_csv(results, out_csv)
    save_xlsx(results, out_xlsx)

    n = len(results)
    print(f"""
╔══════════════════════════════════════════╗
║         QA REPORT — Bolt Food            ║
╠══════════════════════════════════════════╣
║ City:             {city:<22}║
║ Total scraped:    {n:<22}║
║ Missing cuisine:  {sum(1 for r in results if not r.get('cuisine')):<22}║
║ Missing address:  {sum(1 for r in results if not r.get('address')):<22}║
║ Missing rating:   {sum(1 for r in results if not r.get('rating')):<22}║
║ Missing URL:      {sum(1 for r in results if not r.get('restaurant_url')):<22}║
╠══════════════════════════════════════════╣
║ CSV:  {str(out_csv):<36}║
║ XLSX: {str(out_xlsx):<36}║
╚══════════════════════════════════════════╝""")


def main():
    p = argparse.ArgumentParser(description="Bolt Food Playwright scraper")
    p.add_argument("--city",     default="Brno",          help="City name")
    p.add_argument("--max",      type=int, default=100,   help="Max restaurants (default 100)")
    p.add_argument("--headless", action="store_true",     help="Run headless browser")
    p.add_argument("--detail",   action="store_true",     help="Visit each restaurant page (adds cuisine + address)")
    p.add_argument("--lat",      type=float, default=49.1950602, help="City latitude (for unknown cities)")
    p.add_argument("--lng",      type=float, default=16.6068371, help="City longitude (for unknown cities)")
    args = p.parse_args()

    log.info(f"Bolt Food | city={args.city!r} | max={args.max} | headless={args.headless} | detail={args.detail}")
    asyncio.run(run(args.city, args.max, args.headless, args.detail, args.lat, args.lng))


if __name__ == "__main__":
    main()
