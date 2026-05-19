import io
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import WriteOnlyCell

# Wolt-specific columns
WOLT_COLUMNS = [
    ('Restaurant Name',       'name'),
    ('Brand Name',            'brand_name'),
    ('Phone',                 'phone'),
    ('Website',               'website'),
    ('Address',               'address'),
    ('City',                  'city'),
    ('Country',               'country'),
    ('Cuisine / Kitchen',     'cuisine'),
    ('Rating',                'rating'),
    ('Merchant / Legal Co.',  'merchant_name'),
    ('Business ID',           'business_id'),
    ('Legal Street',          'legal_street'),
    ('Legal City',            'legal_city'),
    ('Legal Post Code',       'legal_post_code'),
    ('Legal Country',         'legal_country'),
    ('Platform URL',          'platform_url'),
]

# Bolt-specific columns
BOLT_COLUMNS = [
    ('Restaurant Name',       'name'),
    ('Cuisine / Kitchen',     'cuisine'),
    ('Rating',                'rating'),
    ('Review Count',          'review_count'),
    ('Delivery Fee',          'delivery_fee'),
    ('Delivery Time',         'delivery_time'),
    ('Address',               'address'),
    ('City',                  'city'),
    ('Platform URL',          'platform_url'),
]

# Foodora-specific columns (sitemap only: name + city + country + URL)
FOODORA_COLUMNS = [
    ('Restaurant Name',   'name'),
    ('City',              'city'),
    ('Country',           'country'),
    ('Platform URL',      'platform_url'),
]

# Google Maps columns — clean, minimal set for ChoiceQR prospecting
GOOGLE_MAPS_COLUMNS = [
    ('Name',                  'name'),
    ('Address',               'address'),
    ('City',                  'city'),
    ('Country',               'country'),
    ('Website',               'website'),
    ('Review Count',          'reviews'),
    ('Cuisine / Kitchen',     'category'),
    ('Phone',                 'phone'),
    # Only populated when the restaurant actually has delivery/reservation;
    # empty cell means "none found".
    ('Delivery Companies',    'delivery_companies'),    # e.g. "Wolt, Bolt Food"
    ('Reservation Companies', 'reservation_companies'), # e.g. "Quandoo, TheFork"
]

PLATFORM_COLUMNS = {
    'wolt':         WOLT_COLUMNS,
    'bolt':         BOLT_COLUMNS,
    'foodora':      FOODORA_COLUMNS,
    'google_maps':  GOOGLE_MAPS_COLUMNS,
}

HEADER_COLORS = {
    'wolt':        '009DE0',   # Wolt blue
    'bolt':        '34C759',   # Bolt green
    'foodora':     'E2213D',   # Foodora red
    'google_maps': '34A853',   # Google green
}


def _add_sheet(wb, platform: str, location: str, restaurants: list):
    """Add one sheet to workbook for a given platform (write-only: no cells held in RAM)."""
    columns = PLATFORM_COLUMNS.get(platform, WOLT_COLUMNS)
    color   = HEADER_COLORS.get(platform, '1A73E8')

    sheet_title = f'{platform.capitalize()} - {location}'[:31]
    ws = wb.create_sheet(title=sheet_title)

    # Pre-compute column widths from raw data before writing any rows
    # (must be set before first ws.append() in write-only mode)
    col_widths = [len(header) for header, _ in columns]
    for r in restaurants[:200]:
        for ci, (_, key) in enumerate(columns):
            val_len = len(str(r.get(key, '') or ''))
            if val_len > col_widths[ci]:
                col_widths[ci] = val_len
    for ci, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = min(width + 4, 50)

    hfont  = Font(bold=True, color='FFFFFF')
    hfill  = PatternFill('solid', fgColor=color)
    halign = Alignment(horizontal='center', vertical='center', wrap_text=False)

    header_cells = []
    for header, _ in columns:
        cell = WriteOnlyCell(ws, value=header)
        cell.font, cell.fill, cell.alignment = hfont, hfill, halign
        header_cells.append(cell)
    ws.append(header_cells)

    for r in restaurants:
        ws.append([r.get(key, '') for _, key in columns])


def build_excel(results_by_platform: dict, location: str) -> bytes:
    """Build an Excel file with one sheet per platform (streaming write, low memory)."""
    wb = openpyxl.Workbook(write_only=True)

    for platform, restaurants in results_by_platform.items():
        if restaurants:
            _add_sheet(wb, platform, location, restaurants)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _send(to_email: str, subject: str, html: str, attachment: bytes = None, filename: str = None):
    smtp_host = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
    smtp_port = int(os.environ.get('SMTP_PORT', 587))
    smtp_user = os.environ.get('SMTP_USER')
    smtp_pass = os.environ.get('SMTP_PASSWORD')

    if not smtp_user or not smtp_pass:
        print('[Email] SMTP credentials not set — skipping.')
        return

    msg = MIMEMultipart('mixed')
    msg['Subject'] = subject
    msg['From']    = smtp_user
    msg['To']      = to_email
    msg.attach(MIMEText(html, 'html'))

    if attachment and filename:
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(attachment)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(part)

    with smtplib.SMTP(smtp_host, smtp_port) as server:
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.sendmail(smtp_user, to_email, msg.as_string())

    print(f'[Email] Sent to {to_email}: {subject}')


def send_completion_email(to_email: str, results_by_platform: dict, location: str):
    """
    results_by_platform: {"wolt": [...], "bolt": [...]}
    Builds a single Excel with one tab per platform and emails it.
    """
    excel_bytes = build_excel(results_by_platform, location)

    platforms_str = ' + '.join(p.capitalize() for p in results_by_platform if results_by_platform[p])
    total = sum(len(v) for v in results_by_platform.values())
    safe_loc = location.replace(', ', '_').replace(' ', '_')
    filename = f'{safe_loc}_{platforms_str.replace(" + ", "_")}.xlsx'

    subject = f'✅ Scraping done — {total} restaurants from {location} ({platforms_str})'

    # Build table rows per platform
    platform_rows = ''
    for plat, restaurants in results_by_platform.items():
        if restaurants:
            platform_rows += f"""
        <tr>
          <td style="padding:6px 12px;background:#f4f6fb;font-weight:bold;">{plat.capitalize()}</td>
          <td style="padding:6px 12px;"><strong>{len(restaurants)}</strong> restaurants</td>
        </tr>"""

    html = f"""
    <html><body style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:24px;">
      <h2 style="color:#009de0;">🍽️ Your restaurant data is ready!</h2>
      <p>The scraping job completed successfully. The Excel file is attached.</p>
      <table style="border-collapse:collapse;margin:16px 0;">
        <tr><td style="padding:6px 12px;background:#f4f6fb;font-weight:bold;">Location</td>
            <td style="padding:6px 12px;">{location}</td></tr>
        {platform_rows}
        <tr><td style="padding:6px 12px;background:#f4f6fb;font-weight:bold;">Total</td>
            <td style="padding:6px 12px;"><strong>{total}</strong> restaurants</td></tr>
        <tr><td style="padding:6px 12px;background:#f4f6fb;font-weight:bold;">File</td>
            <td style="padding:6px 12px;">{filename}</td></tr>
      </table>
      <p style="margin-top:24px;color:#6b7280;font-size:0.85rem;">Generated by Restaurant Scraper.</p>
    </body></html>
    """
    _send(to_email, subject, html, attachment=excel_bytes, filename=filename)


def build_google_maps_excel(results: list, location: str, stats: dict = None) -> bytes:
    """
    Build a 4-sheet Excel for Google Maps results:
    1. All Results
    2. Has Delivery  (has delivery_platforms)
    3. Has Reservation (has reservation_system)
    4. Summary stats
    """
    columns = GOOGLE_MAPS_COLUMNS
    color   = '34A853'   # Google green

    wb = openpyxl.Workbook(write_only=True)

    def _write_data_sheet(title, rows):
        ws = wb.create_sheet(title=title[:31])

        # Pre-compute column widths from raw data before first append
        col_widths = [len(h) for h, _ in columns]
        for r in rows[:200]:
            for ci, (_, key) in enumerate(columns):
                col_widths[ci] = max(col_widths[ci], len(str(r.get(key, '') or '')))
        for ci, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = min(w + 4, 60)

        hfont  = Font(bold=True, color='FFFFFF')
        hfill  = PatternFill('solid', fgColor=color)
        halign = Alignment(horizontal='center', vertical='center')

        header_cells = []
        for h, _ in columns:
            c = WriteOnlyCell(ws, value=h)
            c.font, c.fill, c.alignment = hfont, hfill, halign
            header_cells.append(c)
        ws.append(header_cells)

        for r in rows:
            ws.append([r.get(key, '') for _, key in columns])

    # Sheet 1: All Results
    _write_data_sheet('All Results', results)

    # Sheet 2: Has Delivery — restaurants with at least one delivery company
    delivery_rows = [r for r in results if r.get('delivery_companies')]
    _write_data_sheet(f'Has Delivery ({len(delivery_rows)})', delivery_rows)

    # Sheet 3: Has Reservation — restaurants with at least one reservation company
    reservation_rows = [r for r in results if r.get('reservation_companies')]
    _write_data_sheet(f'Has Reservation ({len(reservation_rows)})', reservation_rows)

    # Sheet 4: Summary
    ws_sum = wb.create_sheet(title='Summary')

    from collections import Counter
    delivery_counts: Counter = Counter()
    for r in results:
        for plat in r.get('delivery_companies', '').split(', '):
            if plat.strip():
                delivery_counts[plat.strip()] += 1

    reservation_counts: Counter = Counter()
    for r in results:
        for sys_ in r.get('reservation_companies', '').split(', '):
            if sys_.strip():
                reservation_counts[sys_.strip()] += 1

    city_counts: Counter = Counter(r.get('city', 'Unknown') for r in results)

    hfont = Font(bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor=color)

    def hrow(label):
        c = WriteOnlyCell(ws_sum, value=label)
        c.font, c.fill = hfont, hfill
        return [c]

    ws_sum.column_dimensions['A'].width = 30
    ws_sum.column_dimensions['B'].width = 15

    total = len(results)
    with_phone    = sum(1 for r in results if r.get('phone'))
    with_website  = sum(1 for r in results if r.get('website'))

    def pct(n): return f'{round(n / total * 100)}%' if total else '0%'

    ws_sum.append(hrow(f'Google Maps — {location}'))
    ws_sum.append(['Total found',       total])
    ws_sum.append(['With phone',        with_phone,              pct(with_phone)])
    ws_sum.append(['With website',      with_website,            pct(with_website)])
    ws_sum.append(['With delivery',     len(delivery_rows),      pct(len(delivery_rows))])
    ws_sum.append(['With reservation',  len(reservation_rows),   pct(len(reservation_rows))])
    ws_sum.append([])
    ws_sum.append(hrow('Delivery platforms'))
    for plat, cnt in delivery_counts.most_common():
        ws_sum.append([plat, cnt])
    ws_sum.append([])
    ws_sum.append(hrow('Reservation systems'))
    for sys_, cnt in reservation_counts.most_common():
        ws_sum.append([sys_, cnt])
    ws_sum.append([])
    ws_sum.append(hrow('Top cities'))
    for city, cnt in city_counts.most_common(20):
        ws_sum.append([city, cnt])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_google_maps_completion_email(
    to_email: str, results: list, location: str,
    stats: dict, excel_bytes: bytes, filename: str,
):
    """Send completion email for Google Maps jobs using the pre-built 4-sheet excel."""
    total        = len(results)
    with_phone   = stats.get('with_phone', 0)
    with_website = stats.get('with_website', 0)
    with_del     = stats.get('with_delivery', 0)
    with_res     = stats.get('with_reservation', 0)

    def pct(n): return f'{round(n / total * 100)}%' if total else '0%'

    subject = f'✅ Google Maps scraping done — {total} places from {location}'
    html = f"""
    <html><body style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:24px;">
      <h2 style="color:#34A853;">🗺️ Your Google Maps data is ready!</h2>
      <p>Scraping completed. The Excel file (4 sheets) is attached.</p>
      <table style="border-collapse:collapse;margin:16px 0;width:100%;">
        <tr><td style="padding:6px 12px;background:#f4f6fb;font-weight:bold;">Location</td>
            <td style="padding:6px 12px;">{location}</td></tr>
        <tr><td style="padding:6px 12px;background:#f4f6fb;font-weight:bold;">Total found</td>
            <td style="padding:6px 12px;"><strong>{total:,}</strong></td></tr>
        <tr><td style="padding:6px 12px;background:#f4f6fb;font-weight:bold;">With phone</td>
            <td style="padding:6px 12px;">{with_phone:,} ({pct(with_phone)})</td></tr>
        <tr><td style="padding:6px 12px;background:#f4f6fb;font-weight:bold;">With website</td>
            <td style="padding:6px 12px;">{with_website:,} ({pct(with_website)})</td></tr>
        <tr><td style="padding:6px 12px;background:#f4f6fb;font-weight:bold;">With delivery link</td>
            <td style="padding:6px 12px;">{with_del:,} ({pct(with_del)})</td></tr>
        <tr><td style="padding:6px 12px;background:#f4f6fb;font-weight:bold;">With reservation</td>
            <td style="padding:6px 12px;">{with_res:,} ({pct(with_res)})</td></tr>
        <tr><td style="padding:6px 12px;background:#f4f6fb;font-weight:bold;">File</td>
            <td style="padding:6px 12px;">{filename}</td></tr>
      </table>
      <p style="color:#6b7280;font-size:0.85rem;">Generated by Restaurant Scraper · Google Maps tab</p>
    </body></html>
    """
    _send(to_email, subject, html, attachment=excel_bytes, filename=filename)


# New intel columns appended after original file columns
INTEL_NEW_COLUMNS = [
    ('Instagram URL',         'instagram_url'),
    ('Instagram Followers',   'instagram_followers'),
    ('Facebook URL',          'facebook_url'),
    ('Facebook Followers',    'facebook_followers'),
    ('Email Address(es)',     'emails'),
    ('Legal Name',            'legal_name'),
    ('Company ID',            'company_id'),
    ('IČO',                   'ico'),
    ('Website Platform',      'website_platform'),
    ('Reservation (Y/N)',     'reservation_possible'),
    ('Reservation Provider',  'reservation_provider'),
    ('Online Ordering (Y/N)', 'ordering_possible'),
    ('Ordering Provider',     'ordering_provider'),
    ('Notes',                 'notes'),
]

# Fallback columns when no original file data is present
WEBSITE_INTEL_COLUMNS = [
    ('Restaurant Name',       'name'),
    ('Website URL',           'url'),
] + INTEL_NEW_COLUMNS


def build_website_intel_excel(results: list) -> bytes:
    """Build Excel: original file columns (gray header) + new intel columns (purple header)."""
    PURPLE = '7c3aed'
    GRAY   = '374151'

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title='Website Intelligence')

    hfont  = Font(bold=True, color='FFFFFF')
    halign = Alignment(horizontal='center', vertical='center', wrap_text=False)

    # Detect whether results carry original file data
    first_with_orig = next((r for r in results if r.get('_orig')), None)

    if first_with_orig:
        orig_keys   = list(first_with_orig['_orig'].keys())
        intel_cols  = INTEL_NEW_COLUMNS
        all_headers = orig_keys + [h for h, _ in intel_cols]

        # Pre-compute column widths
        col_widths = [len(h) for h in all_headers]
        for r in results[:200]:
            orig = r.get('_orig', {})
            for ci, k in enumerate(orig_keys):
                col_widths[ci] = max(col_widths[ci], len(str(orig.get(k, '') or '')))
            for ci, (_, key) in enumerate(intel_cols, start=len(orig_keys)):
                col_widths[ci] = max(col_widths[ci], len(str(r.get(key, '') or '')))
        for ci, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = min(w + 4, 60)

        # Header row: gray for original cols, purple for new intel cols
        header_cells = []
        for k in orig_keys:
            c = WriteOnlyCell(ws, value=k)
            c.font  = hfont
            c.fill  = PatternFill('solid', fgColor=GRAY)
            c.alignment = halign
            header_cells.append(c)
        for h, _ in intel_cols:
            c = WriteOnlyCell(ws, value=h)
            c.font  = hfont
            c.fill  = PatternFill('solid', fgColor=PURPLE)
            c.alignment = halign
            header_cells.append(c)
        ws.append(header_cells)

        # Data rows
        for r in results:
            orig = r.get('_orig', {})
            row  = [orig.get(k, '') for k in orig_keys]
            row += [r.get(key, '') for _, key in intel_cols]
            ws.append(row)

    else:
        # No original data — use standard flat column list
        columns = WEBSITE_INTEL_COLUMNS
        col_widths = [len(h) for h, _ in columns]
        for r in results[:200]:
            for ci, (_, key) in enumerate(columns):
                col_widths[ci] = max(col_widths[ci], len(str(r.get(key, '') or '')))
        for ci, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = min(w + 4, 60)

        header_cells = []
        for header, _ in columns:
            c = WriteOnlyCell(ws, value=header)
            c.font  = hfont
            c.fill  = PatternFill('solid', fgColor=PURPLE)
            c.alignment = halign
            header_cells.append(c)
        ws.append(header_cells)

        for r in results:
            ws.append([r.get(key, '') for _, key in columns])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_error_email(to_email: str, platform: str, location: str, error_message: str):
    subject = f'❌ Scraping failed — {platform.capitalize()} / {location}'
    html = f"""
    <html><body style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:24px;">
      <h2 style="color:#ef4444;">❌ Scraping stopped due to an error</h2>
      <table style="border-collapse:collapse;margin:16px 0;">
        <tr><td style="padding:6px 12px;background:#fff5f5;font-weight:bold;">Platform</td>
            <td style="padding:6px 12px;">{platform.capitalize()}</td></tr>
        <tr><td style="padding:6px 12px;background:#fff5f5;font-weight:bold;">Location</td>
            <td style="padding:6px 12px;">{location}</td></tr>
        <tr><td style="padding:6px 12px;background:#fff5f5;font-weight:bold;">Error</td>
            <td style="padding:6px 12px;color:#ef4444;"><code>{error_message}</code></td></tr>
      </table>
      <p style="color:#6b7280;font-size:0.85rem;">Generated by Restaurant Scraper.</p>
    </body></html>
    """
    _send(to_email, subject, html)
